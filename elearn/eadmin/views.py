from io import BytesIO
from django.shortcuts import render, redirect
from django.http import JsonResponse
from django.contrib.auth.hashers import check_password
from .models import Admin, Grade, Section, Subject, SectionSubject, Student, StudentInfo,StudentSubjects, Teacher, TeacherInfo, TeacherSubjects
from datetime import datetime
import json
from .factory import sessionHandeler
from openpyxl import load_workbook
# Create your views here.

def index(request):
    if not sessionHandeler.requestIsValid(request):
        return redirect("/eadmin/login")
    return render(request, 'eadmin/index.html')

# alows admin to create new Grade, add teacher etc.
def create(request):
    if not sessionHandeler.requestIsValid(request):
        return redirect("/eadmin/login")
    tmp = Grade.objects.all()
    allGrade = {f"{i.id}" :f"{i.name}" for i in tmp}
    allInfo = {"grades": allGrade}
    tmp = Section.objects.all()
    allInfo["sections"] = {f"{i.id}" :f"{i.grade.name}: {i.name}" for i in tmp}
    allInfo["teachers"] = {f"{i.username.username}":f"{i.name}" for i in TeacherInfo.objects.all()}
    return render(request, 'eadmin/create.html', context=allInfo)

    
def createGrade(request):
    if not sessionHandeler.requestIsValid(request):
        return redirect("/eadmin/login")
    failedList = []
    if request.method == 'POST':
        for i in request.POST.keys():
            try: 
                if "grade" in i: Grade.objects.create(name=request.POST[i]).save()
            except:
                failedList.append(request.POST[i])
        if bool(failedList):
            return JsonResponse({"message":f"Error occured on adding {str(failedList)}"})
    return render(request, 'eadmin/create.html')

def createSection(request):
    if not sessionHandeler.requestIsValid(request):
        return redirect("/eadmin/login")
    failedList = []
    if request.method == 'POST':
        data = json.loads(request.body)
        for i in data:
            for j in data[i]:
                try:
                    Section.objects.create(grade=Grade.objects.get(id = i), name=j).save()
                except: 
                    failedList.append((Grade.objects.get(id = i).name,j))
        if bool(failedList):
            return JsonResponse({"message":f"Unable to add {str(failedList)}"})
        return JsonResponse({"message": "ok"})
    return JsonResponse({"message":"This should Never Happen"})

def createSubject(request):
    if not sessionHandeler.requestIsValid(request):
        return redirect("/eadmin/login")
    failedList = []
    if request.method == 'POST':
        data = json.loads(request.body)
        for i in data:
            for j in data[i]:
                try:
                    Subject.objects.create(name=j).save()
                    SectionSubject.objects.create(section=Section.objects.get(id = i), subject=Subject.objects.get(name=j)).save()
                except: 
                    failedList.append((Section.objects.get(id = i).name,j))
        if bool(failedList):
            return JsonResponse({"message":f"Unable to add {str(failedList)}"})
        return JsonResponse({"message": "ok"})
    return JsonResponse({"message":"This shoud Never Happen"})

def addStudents(request):
    if not sessionHandeler.requestIsValid(request):
        return redirect("/eadmin/login")
    failedList = []
    if request.method == 'POST':
        excel = request.FILES.get("file")
        sheet = load_workbook(filename=BytesIO(excel.read()))["newStudents"]
        # sheet = workbook.active
        rowCounter = len(sheet['A'])
        for i in range(2,rowCounter+1):
            xlUsername = sheet[f"C{i}"].value
            xlPassword = sheet[f"D{i}"].value
            xlName = sheet[f"B{i}"].value
            xlRoll = sheet[f"A{i}"].value
            xlSection = sheet[f"E{i}"].value
            a = Student.objects.create(username=xlUsername, password=xlPassword)
            a.save()
            StudentInfo.objects.create(username=a, name=xlName, rollNo=xlRoll, section=Section.objects.get(name=xlSection))
        return JsonResponse({"message": "ok"})
    
def addTeachers(request):
    if not sessionHandeler.requestIsValid(request):
        return redirect("/eadmin/login")
    failedList = []
    if request.method == 'POST':
        excel = request.FILES.get("file")
        sheet = load_workbook(filename=BytesIO(excel.read()))["newTeachers"]
        # sheet = workbook.active
        rowCounter = len(sheet['A'])
        for i in range(2,rowCounter+1):
            xlUsername = sheet[f"B{i}"].value
            xlPassword = sheet[f"C{i}"].value
            xlName = sheet[f"A{i}"].value
            a = Teacher.objects.create(username=xlUsername, password=xlPassword)
            a.save()
            TeacherInfo.objects.create(username=a, name=xlName)
        return JsonResponse({"message": "ok"})
    
def assignSubject(request):
    if not sessionHandeler.requestIsValid(request):
        return redirect("/eadmin/login")
    if request.method == "GET":
        username = request.GET.get("assignTeacherSubject")
        json = {"for": username}
        allSubs = {}
        subjects = list(Subject.objects.all())
        for i in subjects:
            grade = (SectionSubject.objects.get(subject = i).section).grade.name
            section = SectionSubject.objects.get(subject = i).section.name
            if grade in allSubs:
                if section in allSubs[grade]: 
                    if TeacherSubjects.objects.filter(username=Teacher.objects.get(username=username), subject=i).exists():
                        (allSubs[grade][section]).append((i.id, i.name, "true"))
                    else: (allSubs[grade][section]).append((i.id, i.name, "false"))
                else:
                    if TeacherSubjects.objects.filter(username=Teacher.objects.get(username=username), subject=i).exists():
                        allSubs[grade][section] = [(i.id, i.name, "true"),]
                    else: allSubs[grade][section] = [(i.id, i.name, "false"),]
            else:
                if TeacherSubjects.objects.filter(username=Teacher.objects.get(username=username), subject=i).exists():
                    allSubs[grade] = {section: [(i.id, i.name, "true"),]}
                else: allSubs[grade] = {section: [(i.id, i.name, "false"),]}
        json["allSubs"] = allSubs
        # return JsonResponse(json)
        return render(request, "eadmin/assignSubject.html", context=json)
    if request.method == "POST":
        teacher = Teacher.objects.get(username=request.POST.get("for"))
        teacherSubjects = list(TeacherSubjects.objects.filter(username=teacher))
        for i in list(request.POST.getlist("subjects")):
            if TeacherSubjects.objects.filter(username=teacher, subject=Subject.objects.get(id = int(i))).exists(): 
                a = TeacherSubjects.objects.get(username=teacher, subject=Subject.objects.get(id = int(i)))
                teacherSubjects.remove(a)
            else: 
                TeacherSubjects.objects.create(username=teacher, subject=Subject.objects.get(id = int(i)))
        if bool(teacherSubjects): 
            for i in teacherSubjects:
                i.delete()
        return redirect("/eadmin/create")

def assignStudentSubject(request):
    if not sessionHandeler.requestIsValid(request):
        return redirect("/eadmin/login")
    if request.method == "GET":
        section = Section.objects.get(id = request.GET.get("section"))
        sectionSubjects =list(SectionSubject.objects.filter(section=section))
        students = list(StudentInfo.objects.filter(section=section))
        json = {
            "for": {
                "sectionID": request.GET.get("section"),
                "sectionName": section.name,
                "className": section.grade.name
            },
            "allSubjects": [k.subject.name for k in sectionSubjects],
            "allStudents": {}
        }
        for i in students:
            studentSubjects = [j.subject for j in list(StudentSubjects.objects.filter(username=i.username))]
            json["allStudents"][i.username.username] = {
                "name": i.name,
                "roll": i.rollNo,
                "subjects": [(k.subject.id, k.subject.name, "true" if k in studentSubjects else "false") for k in sectionSubjects]
            }
        # return JsonResponse(json)
        return render(request, "eadmin/assignStudentSubject.html", context=json)
    if request.method == "POST":
        for i in request.POST:
            if i != 'csrfmiddlewaretoken' and i != 'for':
                subs = list(request.POST.getlist(i))
                print(subs, type(subs))
                thisStudent = Student.objects.get(username=i)
                studentSubjects = [j.subject for j in list(StudentSubjects.objects.filter(username=thisStudent))]
                newSubjects = [Subject.objects.get(id=int(sub)) for sub in subs]
                for j in newSubjects:
                    if StudentSubjects.objects.filter(username=thisStudent, subject=SectionSubject.objects.get(subject=j)).exists():
                        studentSubjects.remove(SectionSubject.objects.get(subject=j))
                    else: StudentSubjects.objects.create(username=thisStudent, subject=SectionSubject.objects.get(subject=j))
                if bool(studentSubjects):
                    for j in studentSubjects:
                        StudentSubjects.objects.get(username=thisStudent, subject=j).delete()
        return redirect("/eadmin/create")

        



def login(request):
    if request.method == 'POST':
        uname = request.POST.get('username')
        pswd = request.POST.get('password')
        try:
            userInQuestion = Admin.objects.get(username=uname)
            if check_password(pswd, userInQuestion.password) and userInQuestion.AdminInfo.accessLevel == "SuperUser":
                request.session["isLoggedIn"] = True
                request.session["username"] = userInQuestion.username
                request.session["name"] = userInQuestion.AdminInfo.name
                request.session["loggedInTime"] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                request.session["activeTime"] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                request.session["accessLevel"] = "Admin"
                # return JsonResponse({"message":"OK"})
                return redirect("/eadmin/index")
            return JsonResponse({"message":"Authorization Failed"})
        except Admin.DoesNotExist:
            return JsonResponse({"message":"Invalid Username"})
    return render(request, 'eadmin/login.html')
